from django.conf.urls import url
from django.template.response import TemplateResponse
from resource_tracking.models.tracking_summary import TrackingSummary
from django.contrib import admin

# Register your models here.

from siteapp.settings import BASE_DIR
from sites.models.sites import Sites
from django import forms
from django.contrib import messages
from django.db.models.query_utils import Q
from django.db import models


from django.http.response import HttpResponseRedirect
from typing import Set
from django.contrib.auth.models import User
from django.shortcuts import render
from django.utils.html import format_html
from django.urls import reverse
from datetime import timedelta, datetime, timezone,date
from django.urls import reverse
from django.utils.http import urlencode
from django.http import HttpResponse
from django.apps import apps
from resource_tracking.models.attendance import Attendance
from resource_tracking.models.fault_management import FaultManagement
from resource_tracking.models.faultsite_images import FaultImages
from resource_tracking.models.site_tracking import SiteTracking
from django.db.models.signals import post_save
from django.dispatch import receiver
from resource_tracking.models.site_assignment import SiteAssignement
from resource_tracking.models.fault_assignment import FaultAssignment
from resource_tracking.models.faultsite_comments import FaultComments


from django.db.models import Sum, Count, Avg, Q
import json
from push_notifications.models import GCMDevice
import requests
from django.conf import settings
from rangefilter.filters import DateRangeFilter, DateTimeRangeFilter
import openpyxl
from resource_tracking.models.push_notification_log import PushNotificationLog
from django.contrib.admin.views.main import ChangeList


# Register your models here.
class FaultImagesInline(admin.TabularInline):
    model = FaultImages
    # extra=0
    def get_extra(self, request, obj=None, **kwargs):
        if obj :
            return 1
        else:
            return 0
    fieldsets = [
        (None, {'fields': [('image_name')]}),
    ]

class FaultCommentsInline(admin.TabularInline):
    model = FaultComments
    # extra=0
    def get_extra(self, request, obj=None, **kwargs):
        if obj :
            return 1
        else:
            return 0
    fieldsets = [
        (None, {'fields': [('created_by','comments','created_at')]}),
    ]

class AdminFaultManagement(admin.ModelAdmin):
    model = FaultManagement
    list_display =['id','site','status','fault_description', 'created_at','track_user', 'tracking_summary']

    inlines = [FaultImagesInline,FaultCommentsInline]
    # filter_horizontal=['user']


    fieldsets = [
        ('', {'fields': ['site']}),
        ('', {'fields': ['user']}),
        ('', {'fields': ['fault_description']}),
        ('', {'fields': [('status')]}),
        ('', {'fields': [('admin_remark')]}),
    ]


    def get_form(self, request, obj=None, **kwargs):
        form = super(AdminFaultManagement, self).get_form(request, obj, **kwargs)
        form.base_fields['site'].label_from_instance = lambda obj: "{}".format(obj.site_name)
        return form
    
    def track_user(self,instance):
        ticket_id=(instance)
        # print(ticket_id)        
        return format_html(f'''<a href='/resource_tracking/track_user/?ticket_id={ticket_id}' style="padding: 5px ;" target="_blank" rel="noopener noreferrer">View<a/>''')
    
    def tracking_summary(self, instance):
        ticket_id=(instance)
        return format_html(f'''<a href='/resource_tracking/track_summary/?ticket_id={ticket_id}' style="padding: 5px ;" target="_blank" rel="noopener noreferrer">View<a/>''')
    
    def save_model(self, request, obj, form, change):
        # print("here in save ==========================", request.POST)
        # print()
        if getattr(obj, 'created_by', None) is None:
            # print("USER IS NOT PRESSENT ************")
            # print()
            obj.created_by = request.user
            obj.modified_by = request.user
            
        if  obj.status == "Approved":
            obj.approved_by = request.user

        obj.save()
        super().save_model(request, obj, form, change)

    def save_related(self, request, form, formsets, change):
        # print("in super ")
        super(AdminFaultManagement, self).save_related(request, form, formsets, change)
        # print("after super ")
        obj = form.instance
        obj.save()
        # send the push notification
        # print("Object",obj.id)
        # print("Form change",change)

        # if change==False:
        #     FultData=FaultManagement.objects.filter(id=obj.id)
        #     for fault in FultData:
        #          for rolename in fault.user.all():
        #             useList=User.objects.filter(id=rolename.id)
        #             for user in useList:
        #                 print("User Id",user.id)
        #                 pushResponse = send_push_notification(user,fault)
        #                 # print(pushResponse)
        #                 if pushResponse:
        #                     pushObject = PushNotificationLog()
        #                     pushObject.user = user
        #                     pushObject.fault = fault
        #                     pushObject.notification_response = pushResponse
        #                     pushObject.created_by = request.user
        #                     pushObject.modified_by = request.user
        #                     pushObject.save()
                                               
    class Media:
        css = {
                "all": ("resource_tracking/css/floatLeftCss.css",)
            }
        
def send_push_notification(user,fault):
    try:
        device = GCMDevice.objects.get(user=user)
        if device:
            deviceToken=device.registration_id
            # print("Token",deviceToken)
            serverToken =settings.PUSH_NOTIFICATIONS_SETTINGS['FCM_API_KEY']
            # print("Server",serverToken)
            headers = {
                    'Content-Type': 'application/json',
                    'Authorization': 'key=' + serverToken,
                }
            body = {
                    'data': {'title': 'Fault Assignment',
                                        'body': fault.site.site_name+' assign to you'
                                        },
                    'to':
                        deviceToken,
                    'priority': 'high',
                    #   'data': dataPayLoad,
                    }


            response = requests.post("https://fcm.googleapis.com/fcm/send",headers = headers, data=json.dumps(body))
            # check what is the response
            # print(response)
            
            return response
    except Exception as e:
        print(e) 

class TotalAveragesChangeList(ChangeList):
    #provide the list of fields that we need to calculate averages and totals
    fields_to_total = ['total_time', 'distance',]

class AdminSiteTracking(admin.ModelAdmin):

    list_display =['ticket', 'created_by','status',"status_comment","latitude","longitude", "total_time", "distance","address","created_at"]
    # list_filter = (
    #     ('created_at', DateRangeFilter),
    #     'created_by'
    # )
    def get_model_perms(self, request):
        """
        Return empty perms dict thus hiding the model from admin index.
        """
        return {}
    def get_changelist(self, request, **kwargs):
        return TotalAveragesChangeList

# @admin.register(TrackingSummary)
class AdminTrackingSummary(admin.ModelAdmin):
    list_display =['ticket', 'created_by','distance',"total_time"]
    # list_filter = (
    #     ('created_at', DateRangeFilter),
    #     'created_by'
    # )
    search_fields=['ticket__id']

    change_list_template = 'admin/tracking_summary_change_list.html'
    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(
            request,
            extra_context=extra_context,
        )
        try:
            qs = response.context_data['cl'].queryset
        except (AttributeError, KeyError):
            return response

        metrics = {
            'distance': Sum('distance'),
            'total_time':Sum('total_time')/60,
        }
        response.context_data['summary'] = list(
            qs
            .values('ticket', 'ticket__site__site_name')
            .annotate(**metrics)
            .order_by('-distance')
        )

        return response

    def get_urls(self):
        urls = super(AdminTrackingSummary, self).get_urls()
        my_urls = [
            url('get_report/', self.get_report)
        ]
        return my_urls + urls

    def get_report(self, request):
        # ...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
        context = dict(
           # Include common variables for rendering the admin template.
           self.admin_site.each_context(request),
           # Anything else you want in the context...
           key=1,
        )
        return TemplateResponse(request, "sitetracking/sometemplate.html", context)

class AdminSiteAssignment(admin.ModelAdmin):
    list_display =['site_oprator_id','site']
    filter_horizontal=['user']
    fieldsets = [
        ('', {'fields': ['site']}),
        ('', {'fields': ['user']}),
    ]
    def site_oprator_id(self, instance):
        siteObj=Sites.objects.get(id=instance.site.id)
        return siteObj.site_operator_id
    def get_form(self, request, obj=None, **kwargs):
        form = super(AdminSiteAssignment, self).get_form(request, obj, **kwargs)
        form.base_fields['site'].label_from_instance = lambda obj: "{} ({})".format(obj.site_operator_id, obj.site_name)
        return form
    def save_model(self, request, obj, form, change):
        if getattr(obj, 'created_by', None) is None:
            obj.created_by = request.user
            obj.modified_by = request.user
            obj.save()
        super().save_model(request, obj, form, change)
    class Media:
        css = {
                "all": ("resource_tracking/css/floatLeftCss.css",)
            }

class AdminFaultAssignment(admin.ModelAdmin):
    list_display =["site_data"]
    fieldsets = [
        ('', {'fields': ['site_data']}),
    ]
    def save_model(self, request, obj, form, change):
        if getattr(obj, 'created_by', None) is None:
            obj.created_by = request.user
            obj.modified_by = request.user
            obj.save()
            try:
                # static declarations
                pushResponse = ''
                user = ''
                mappedSites = ''
                notMappedSites = ''
                excel_file = request.FILES["site_data"]
                wb = openpyxl.load_workbook(excel_file)

                # getting a particular sheet by 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name out of many sheets
                worksheet = wb["Sheet1"]
                column_name = 'SITEID'
                for column_cell in worksheet.iter_cols(1, worksheet.max_column):
                    if column_cell[0].value == column_name:
                        for data in column_cell[1:]:    # iterate your column
                            # print(data.value)
                            siteObject = Sites.objects.filter(site_operator_id=str(data.value))
                            if siteObject:
                                result = checkSiteValidations(siteObject)
                                if result == False:
                                    continue
                                siteAssignObject = SiteAssignement.objects.filter(site__in=siteObject)
                                if siteAssignObject:
                                    for sitedata in siteAssignObject:
                                        site_users = sitedata.user.all()
                                        faultObject = FaultManagement()
                                        faultObject.site = sitedata.site
                                        faultObject.created_by = request.user
                                        faultObject.modified_by = request.user
                                        faultObject.save()

                                        for user in site_users:
                                            faultObject.user.add(user)
                                            faultObject.save()

                                            pushResponse = send_push_notification(user,faultObject)
                                            pushObject = PushNotificationLog()
                                            pushObject.user = user
                                            pushObject.fault = faultObject
                                            pushObject.notification_response = pushResponse
                                            pushObject.created_by = request.user
                                            pushObject.modified_by = request.user
                                            pushObject.save()

                                        # SLA started
                                        sitetrack = SiteTracking()
                                        sitetrack.ticket = faultObject
                                        sitetrack.latitude = 0
                                        sitetrack.longitude = 0
                                        sitetrack.created_by = request.user
                                        sitetrack.modified_by = request.user
                                        sitetrack.distance = 0
                                        sitetrack.status = "Fault Assigned"
                                        sitetrack.address=""
                                        sitetrack.total_time=0
                                        sitetrack.save()

                                    mappedSites = str(mappedSites) +  ", "+  str(data.value)

                            else:
                                notMappedSites = str(notMappedSites) +  ", "+  str(data.value)

                        break
                succeedMessage = "These sites are mapped successfully - "+ str(mappedSites)
                failedMessage =  "These sites are not mapped - "+ str(notMappedSites)
                messages.add_message(request, messages.INFO, succeedMessage)
                messages.add_message(request, messages.WARNING, failedMessage)
                # import pdb;pdb.set_trace()

            except Exception as e:
                   print(e)
                  
    class Media:
        css = {
                "all": ("resource_tracking/css/floatLeftCss.css",)
            }

def checkSiteValidations(siteObject):
    if siteObject:
        for site in siteObject:
            faultData = FaultManagement.objects.filter(site=site)
            if faultData:
                for fault in faultData:
                    if fault.status=="Approved" and fault.fault_type!="Active Issue":
                        # print("faultData:", faultData)
                        return True
                    else:
                       return False
            return True

class UserTrackingSummary(admin.ModelAdmin):
    pass

# admin.site.register(FaultManagement, AdminFaultManagement)
# admin.site.register(SiteTracking, AdminSiteTracking) 
# admin.site.register(SiteAssignement, AdminSiteAssignment)
# admin.site.register(FaultAssignment, AdminFaultAssignment)


